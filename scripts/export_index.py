from __future__ import annotations

import csv
from pathlib import Path


REPO_ROOT = Path(__file__).resolve().parents[1]
QUEUE_PATH = REPO_ROOT / "queue" / "words.csv"
EXPORTS_DIR = REPO_ROOT / "exports"
CSV_PATH = EXPORTS_DIR / "dictionary_index.csv"
XLSX_PATH = EXPORTS_DIR / "dictionary_index.xlsx"
OUTPUT_COLUMNS = [
    "headword",
    "type",
    "status",
    "priority",
    "file",
    "prompt_version",
    "checked",
    "updated_at",
    "notes",
    "file_exists",
]


def _resolve_entry_path(file_value: str) -> Path | None:
    if not file_value:
        return None
    path = Path(file_value)
    if not path.is_absolute():
        path = REPO_ROOT / path
    return path


def _entry_exists(file_value: str) -> bool:
    path = _resolve_entry_path(file_value)
    if path is None:
        return False
    return path.is_file()


def _hyperlink_target(file_value: str) -> str | None:
    path = _resolve_entry_path(file_value)
    if path is None or not path.is_file():
        return None
    return str(path.resolve())


def _read_queue() -> list[dict[str, str]]:
    with QUEUE_PATH.open("r", encoding="utf-8", newline="") as handle:
        return list(csv.DictReader(handle))


def _build_rows(queue_rows: list[dict[str, str]]) -> list[dict[str, str]]:
    rows: list[dict[str, str]] = []
    for row in queue_rows:
        output = {column: row.get(column, "") or "" for column in OUTPUT_COLUMNS if column != "file_exists"}
        output["file_exists"] = "true" if _entry_exists(row.get("file", "") or "") else "false"
        rows.append(output)
    return rows


def _write_csv(rows: list[dict[str, str]]) -> None:
    EXPORTS_DIR.mkdir(parents=True, exist_ok=True)
    with CSV_PATH.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=OUTPUT_COLUMNS)
        writer.writeheader()
        writer.writerows(rows)


def _write_xlsx(rows: list[dict[str, str]]) -> bool:
    try:
        from openpyxl import Workbook
    except ImportError:
        return False

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "dictionary_index"
    sheet.append(OUTPUT_COLUMNS)
    for row in rows:
        sheet.append([row.get(column, "") for column in OUTPUT_COLUMNS])
        file_cell = sheet.cell(row=sheet.max_row, column=OUTPUT_COLUMNS.index("file") + 1)
        target = _hyperlink_target(row.get("file", ""))
        if target:
            file_cell.hyperlink = target
            file_cell.style = "Hyperlink"
    for column_cells in sheet.columns:
        width = max(len(str(cell.value or "")) for cell in column_cells)
        sheet.column_dimensions[column_cells[0].column_letter].width = min(max(width + 2, 10), 60)
    workbook.save(XLSX_PATH)
    return True


def main() -> int:
    if not QUEUE_PATH.is_file():
        print(f"ERROR: queue file not found: {QUEUE_PATH}")
        return 1
    rows = _build_rows(_read_queue())
    _write_csv(rows)
    print(f"Wrote {CSV_PATH}")
    if _write_xlsx(rows):
        print(f"Wrote {XLSX_PATH}")
    else:
        print("openpyxl is not installed; skipped Excel export")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
